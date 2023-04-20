import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from instagram.client import InstagramAPI

nomes_urna = []
cidade = []
resultados = []
resultados2 = []

api = InstagramAPI(client_id='sidneyoliveira0', client_secret='99832642s')

workbook = openpyxl.load_workbook("padraos/prefeitos.xlsx")

# seleciona a planilha ativa
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=4, max_col=4, values_only=True):
    nomes_urna.append(row[3])
    cidade.append(row[0])

for i in range(len(cidade)):

    url = f"https://www.google.com/search?q={nomes_urna[i]}+prefeito+{cidade[i]}+instagram&lr=lang_pt/"
    print(url)
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    result = str(soup.find("div", class_="BNeawe UPmit AP7Wnd lRVwie"))
    print(result)
    result = result.split(">")
    result = result[1].split("›")

    print(result)
    if result[0] == "www.instagram.com ":
        result = result[1].split("<")
        print(result)
        if len(result) >= 2 and result[0]:
            result = result[0]
        print(result)
        if result:
            resultados.append(result)
            print(f'resultados+{resultados}')
            resultados[i] = resultados[i].replace(" ","")
            worksheet.cell(row=i+4, column=8, value=resultados[i])
            print(resultados)
            workbook.save("prefeitos_do_ceara.xlsx")

            result = str(soup.find("div", class_="BNeawe s3v9rd AP7Wnd"))
            result = result.lower()
            result = result.split("followers")
            print(result)
            result = result[0].split(" ")
            print(result)
            print(result[-2])

            if len(result) >= 2 and result[-2]:
                result = result[-2]
                result = result.replace('ap7wnd">','')
                print(result)
                if result:
                    resultados2.append(result)
                    resultados2[i] = resultados2[i].replace(" ", "")
                    worksheet.cell(row=i + 4, column=9, value=resultados2[i])
                    print(resultados2)

            else:
                resultados.append(" ")
                resultados2.append(" ")
                worksheet.cell(row=i + 4, column=9, value='0')

                print(f"Nenhum resultado encontrado 1")

    else:
        result = str(soup.find("div", class_="BNeawe s3v9rd AP7Wnd"))
        print(result)
        result = result.split(">")
        result = result[1].split("›")

        print(result)
        if result[0] == "www.instagram.com ":
            result = result[1].split("<")
            if len(result) >= 2 and result[0]:
                result = result[0]
            print(result)
            if result:
                resultados.append(result)
                print(f'resultados+{resultados}')
                resultados[i] = resultados[i].replace(" ", "")
                worksheet.cell(row=i + 4, column=8, value=resultados[i])
                print(resultados)

                result = str(soup.find("div", class_="BNeawe s3v9rd AP7Wnd"))

                result = result.split("Followers")
                print(result[0])
                result = result[0].split(" ")
                print(result)
                print(result[-2])

                if len(result) >= 2 and result[-2]:
                    result = result[-2]
                    print(result)
                    if result:
                        resultados2.append(result)
                        resultados2[i] = resultados2[i].replace(" ", "")
                        worksheet.cell(row=i + 4, column=9, value=resultados2[i])
                        print(resultados2)
                        workbook.save("prefeitos_do_ceara.xlsx")

                else:
                    resultados.append(" ")
                    resultados2.append(" ")
                    worksheet.cell(row=i + 4, column=9, value='0')
                    print(f"Nenhum resultado encontrado 1")

        else:
            print(f"Sem insta")
            resultados.append(" ")
            resultados2.append(" ")


            worksheet.cell(row=i + 2, column=9, value='0')

workbook.save("prefeitos_do_ceara.xlsx")
workbook.close()



