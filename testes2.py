import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from instagram.client import InstagramAPI

nomes_urna = []
cidade = []
resultados = []
resultados2 = []
resultados3 = []

api = InstagramAPI(client_id='sidneyoliveira0', client_secret='99832642s')

workbook = openpyxl.load_workbook("padraos/prefeitos.xlsx")

# seleciona a planilha ativa
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=4, max_col=4, values_only=True):
    nomes_urna.append(row[3])
    cidade.append(row[0])

for i in range(len(cidade)):

    url = f"https://www.google.com/search?q={nomes_urna[i]}+prefeito+{cidade[i]}+instagram&lr=lang_pt/"
    # print(url)
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    result = str(soup.find("div", class_="BNeawe UPmit AP7Wnd lRVwie"))
    # print(result)
    result = result.split(">")
    result = result[1].split(" › ")

    # print(result)
    if result[0] == "www.instagram.com":
        result = result[1].split("<")
        print(result)
        if len(result) >= 2 and result[0]:
            result = result[0]
            print(result)
        if result:
            resultados.append(result)
            worksheet.cell(row=i+4, column=8, value=resultados[i])
            print(resultados[i])

            url = f"https://www.instagram.com/{resultados[i]}/?hl=pt"
            # print(url)
            page = requests.get(url)
            soup = BeautifulSoup(page.content, 'html.parser')
            # print(soup.prettify())
            result = str(soup.find("meta", attrs={"name": "description"})).lower()
            # print(result)
            result = result.split(" seguidores")
            result = result[0].replace('<meta content="','')

            resultados2.append(result)
            resultados2[i] = resultados2[i].replace(" ", "")
            resultados2[i] = resultados2[i].replace(",", "")
            resultados2[i] = resultados2[i].replace(".", "")
            resultados2[i] = resultados2[i].replace("k", "000")
            print(resultados2[i])

            if resultados2[i] == 'none' or not resultados2[i].isdigit():
                resultados2[i] = 0
            worksheet.cell(row=i+4, column=9, value=int(resultados2[i]))
            workbook.save("prefeitos_do_ceara.xlsx")
            resultados3.append(" ")

    else:

        if result[0] == "www.facebook.com":
            result = result[1].split("<")
            # print(result)
            if len(result) >= 2 and result[0]:
                result = result[0]
            # print(result)
            if result:
                resultados.append(result)
                # print(f'resultados+{resultados}')
                worksheet.cell(row=i + 4, column=8, value=resultados[i])
                print(resultados[i])

                url = f"https://www.facebook.com/{resultados[i]}"
                # print(url)
                page = requests.get(url)
                soup = BeautifulSoup(page.content, 'html.parser')
                # print(soup.prettify())
                result = str(soup.find("meta", attrs={"name": "description"})).lower()
                # print(result)
                result = result.split(" curtidas")
                # print(result)
                result = result[0].split(" ")
                result = str(result[-1])
                # print(result[-1])
                resultados3.append(result)

                resultados3[i] = resultados3[i].replace(" ", "")
                resultados3[i] = resultados3[i].replace(",", "")
                resultados3[i] = resultados3[i].replace(".", "")
                resultados3[i] = resultados3[i].replace("k", "000")

                if resultados3[i] == 'none' or isinstance(resultados3[i], int):
                    resultados3[i] = 0

                print(resultados3[i])
                worksheet.cell(row=i + 4, column=9, value=int(resultados3[i]))
                workbook.save("prefeitos_do_ceara.xlsx")
                resultados2.append(" ")
        else:
            resultados.append(" ")
            worksheet.cell(row=i + 4, column=8, value='0')
            resultados2.append(" ")
            worksheet.cell(row=i + 4, column=9, value='0')
            resultados3.append(" ")



workbook.save("prefeitos_do_ceara.xlsx")
workbook.close()



