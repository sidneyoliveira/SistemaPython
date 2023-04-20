import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from instagram.client import InstagramAPI

nomes_urna = []
cidade = []
resultados = []
resultados2 = []

workbook = openpyxl.load_workbook("padraos/prefeitos.xlsx")

# seleciona a planilha ativa
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=4, max_col=4, values_only=True):
    nomes_urna.append(row[3])
    cidade.append(row[0])

for i in range(len(cidade)):

    url = f"https://www.google.com/search?q={nomes_urna[i]}+prefeito+{cidade[i]}+instagram&lr=lang_pt/"
    print(url)
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    result = str(soup.find("div", class_="BNeawe UPmit AP7Wnd lRVwie"))
    print(result)
    result = result.split(">")
    result = result[1].split("›")

    print(result)
    if result[0] == "www.instagram.com ":
        result = result[1].split("<")
        print(result)
        if len(result) >= 2 and result[0]:
            result = result[0]
        print(result)
        if result:
            resultados.append(result)
            print(f'resultados+{resultados}')
            resultados[i] = resultados[i].replace(" ","")
            worksheet.cell(row=i+4, column=8, value=resultados[i])
            print(resultados)

            url = f"https://www.facebook.com/{resultados[i]}"
            print(url)
            page = requests.get(url)
            soup = BeautifulSoup(page.content, 'html.parser')
            # print(soup.prettify())
            result = str(soup.find("meta", attrs={"name": "description"})).lower()
            print(result)
            result = result.split(" seguidores")
            result = result[0].replace('<meta content="', '')
            print(result)

    else:
        resultados.append(" ")
        worksheet.cell(row=i + 4, column=8, value='0')
        resultados2.append(" ")
        worksheet.cell(row=i + 4, column=9, value='0')


workbook.save("prefeitos_do_ceara.xlsx")
workbook.close()



