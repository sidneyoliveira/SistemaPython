import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


nomes_urna = []
cidade = []
usuario = []
resultados = []
resultados2 = []
resultados3 = []

workbook = openpyxl.load_workbook("padraos/prefeitos.xlsx")

# seleciona a planilha ativa
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=4, max_col=9, values_only=True):
    nomes_urna.append(row[3])
    cidade.append(row[0])
    usuario.append(row[8])

driver = webdriver.Chrome()
# acessa a URL
url = f"https://www.tucktools.com/instagram-live-followers/"
driver.get(url)
driver.maximize_window()

for i in range(len(usuario)):

    driver.execute_script('window.scrollBy(0, 100);')
    input_element = driver.find_element('id', 'ig-input')
    input_element.clear()  # limpar o valor atual
    input_element.send_keys(usuario[i])
    time.sleep(2)
    button_element = driver.find_element(By.CLASS_NAME, 'btn-primary')
    # imprima o texto
    print(button_element)
    button_element.click()

    time.sleep(10)
    driver.execute_script('window.scrollBy(0, 400);')
    print(driver.find_element(By.CLASS_NAME, 'ml-3').text)

    odometros = driver.find_elements(By.CLASS_NAME, 'odometer-value')
    odometros_texto = []

    for odometro in odometros:
        print(odometro.text)
        odometros_texto.append(odometro.text)

    odometros_string = ''.join(odometros_texto)
    print(odometros_string)

    worksheet.cell(row=i + 4, column=10, value=int(odometros_string))
    workbook.save("prefeitos_do_ceara.xlsx")
    resultados3.append(" ")
    time.sleep(4)


# driver.quit()
workbook.save("prefeitos_do_ceara2.xlsx")
workbook.close()



